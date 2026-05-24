from datetime import timedelta
import random
from io import BytesIO

from django.core.exceptions import ObjectDoesNotExist
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from rest_framework import viewsets, permissions, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.views import APIView

from .models import Center, Subject, Level, Student, Question, Result, StudentAnswer, MentalTask

def is_mental_subject(subject_name):
    normalized = subject_name.lower().replace("'", '').replace('‘', '').replace('’', '')
    return 'mental' in normalized


def get_exam_duration_minutes(student):
    if is_mental_subject(student.subject.name):
        return int(getattr(student.level, 'duration_minutes', 10) or 10)
    return int(getattr(student.level, 'duration_minutes', 30) or 30)

def exam_duration_seconds(student):
    return max(1, get_exam_duration_minutes(student) * 60)


def get_remaining_seconds(student, now=None):
    """Backend qolgan vaqtni o'zi hisoblaydi. Frontend vaqt zonasi yoki
    browser Date.parse xatolari sabab 00:00 bo'lib qolmasligi uchun shu qiymat
    exam/start javobida yuboriladi.
    """
    now = now or timezone.now()
    duration_seconds = exam_duration_seconds(student)
    if not student.started_at:
        return duration_seconds
    elapsed = int((now - student.started_at).total_seconds())
    if elapsed < 0:
        elapsed = 0
    return max(0, duration_seconds - elapsed)


def calculate_spent_seconds_from_request(request, student, finished_at):
    """Natijadagi vaqt code kiritilgan paytdan hisoblanadi.
    Frontend yuborgan elapsed qiymat noto'g'ri bo'lsa, serverdagi started_at asosiy manba bo'ladi.
    """
    duration_seconds = exam_duration_seconds(student)
    raw_elapsed = request.data.get('client_elapsed_seconds')
    try:
        client_elapsed = int(float(raw_elapsed))
    except (TypeError, ValueError):
        client_elapsed = None

    if client_elapsed is not None and 0 <= client_elapsed <= duration_seconds:
        return client_elapsed

    started_at = student.started_at or finished_at
    return max(0, min(int((finished_at - started_at).total_seconds()), duration_seconds))



MENTAL_5_6_YOSH_SEQUENCES = [[5, 1, 3, -4],
 [2, 1, 5, -8],
 [5, 4, -3, -1],
 [9, -4, -5, 6],
 [2, 7, -6, 5],
 [8, -1, -5, 6],
 [2, 1, -3, 4],
 [4, -2, -1, 3],
 [9, -2, -7, 8],
 [2, 1, 5, -3],
 [7, 2, -4, -5],
 [3, 5, 1, -3],
 [7, -2, 1, -5],
 [4, -2, -1, 6],
 [3, 5, 1, -3],
 [6, 3, -9, 2],
 [4, 5, -3, -5],
 [5, 3, -2, 3],
 [7, -5, 7, -9],
 [4, -2, -1, 8],
 [9, -1, -5, 1],
 [1, 3, -4, 9],
 [4, -1, 5, -7],
 [3, 6, -3, 1],
 [7, 1, -3, 1],
 [1, 7, 1, -7],
 [9, -2, -2, 1],
 [5, 3, -2, 1],
 [8, -1, -6, 7],
 [9, -5, -2, 1],
 [5, 3, -5, 1],
 [9, -3, -6, 1],
 [5, 4, -3, -1],
 [4, -1, -3, 4],
 [9, -1, -7, 5],
 [1, 3, -2, 7],
 [2, 7, -9, 2],
 [8, -3, -5, 9],
 [7, -1, 3, -5],
 [6, -5, 3, -2],
 [2, 6, -3, 2],
 [5, 3, -5, 6],
 [4, -1, -3, 8],
 [3, -2, 7, -3],
 [2, -1, 5, 1],
 [5, 1, 3, -6],
 [8, -7, 2, -1],
 [1, 5, 1, -6],
 [3, 5, 1, -7],
 [2, -1, 2, 1],
 [3, 5, 1, -2],
 [1, 3, -2, -1],
 [8, -6, 1, -2],
 [4, 5, -6, 5],
 [9, -6, 1, -2],
 [7, 1, -5, -3],
 [1, 5, -1, 3],
 [6, -1, 2, -7],
 [3, -1, 7, -8],
 [8, -3, 1, -5],
 [6, -5, 3, -2],
 [2, 6, -5, 6],
 [8, -3, 1, -6],
 [7, -5, -1, 6],
 [6, 2, 1, -8],
 [9, -5, -3, -1],
 [1, 8, -9, 5],
 [5, 3, -2, -6],
 [9, -4, 3, -5],
 [8, -7, 6, -2],
 [4, -2, 6, -8],
 [7, 1, -5, -1],
 [9, -2, -7, 8],
 [2, -1, 8, -9],
 [7, -1, 3, -1],
 [5, 4, -2, -6],
 [7, -5, -1, 8],
 [3, -2, 7, -2],
 [5, 3, -8, 2],
 [4, -2, -2, 7],
 [6, -1, 2, -6],
 [4, 5, -8, -1],
 [7, -6, 1, 5],
 [5, 2, 2, -5],
 [9, -3, -6, 5],
 [5, 1, 3, -6],
 [9, -3, -6, 5],
 [8, -3, 4, -2],
 [4, 5, -9, 2],
 [1, 2, -3, 1],
 [5, 3, 1, -2],
 [8, -3, -5, 2],
 [4, -2, -2, 7],
 [7, -2, 3, -7],
 [4, 5, -7, 5],
 [7, 1, -7, -1],
 [5, 2, -1, -6],
 [8, -3, 2, -6],
 [2, 6, -2, 1],
 [7, -1, 2, -5]]

MENTAL_8_9_YOSH_SEQUENCES = [[4, 5, -6, 5, -2, 3],
 [1, 5, 3, -4, 3, -7],
 [4, -2, 5, 1, -5, -1],
 [6, 1, -2, 4, -5, -1],
 [2, 2, 5, -9, 5, 4],
 [9, -1, -2, -1, 3, -7],
 [1, 3, 5, -7, -1, 5],
 [7, -5, -1, 5, 3, -1],
 [3, -2, 5, 2, -1, -7],
 [2, 6, -8, 9, -2, -5],
 [3, 5, -2, -5, 3, -1],
 [4, -2, 5, -7, 5, 3],
 [5, 4, -9, 7, -2, -5],
 [3, -2, 6, -2, 4, -3],
 [9, -8, 5, -1, -5, 3],
 [8, -1, -7, 1, 6, -2],
 [6, 3, -5, -4, 1, 6],
 [3, 1, 5, -6, -2, 1],
 [2, -1, 6, -7, 9, -1],
 [7, 1, -2, -5, 7, -5],
 [4, -2, 6, 1, -3, -1],
 [5, 1, 2, -7, 8, -9],
 [9, -2, -6, -1, 6, -5],
 [6, 1, 2, -6, -3, 5],
 [8, -2, -5, 7, -2, -5],
 [2, 1, -3, 6, 2, -1],
 [3, -2, 5, 3, -4, -5],
 [5, 1, 1, -2, 1, 2],
 [7, -1, -6, 2, 5, -1],
 [6, 3, -8, 3, -4, 8],
 [7, -6, 8, -7, -2, 9],
 [1, 5, 2, -7, 5, 2],
 [4, -1, 6, -4, -5, 6],
 [3, 6, -9, 6, 1, -2],
 [8, 1, -7, 5, -6, 3],
 [3, 5, -6, 2, -4, 3],
 [9, -8, -1, 2, -1, 8],
 [4, 5, -7, 2, 5, -9],
 [7, 2, -4, 2, 2, -5],
 [2, 1, 1, -2, 5, -6],
 [3, 6, -7, -1, 6, -2],
 [9, -3, -6, 9, -6, -2],
 [3, -2, 5, -6, 5, 3],
 [1, 2, 1, 5, -7, 5],
 [4, 5, -9, 2, 7, -5],
 [2, 5, -2, 4, -1, -6],
 [4, -1, 5, -1, -7, 6],
 [7, 2, -7, -2, 8, -5],
 [2, 6, -7, 5, 2, -3],
 [5, 2, -7, 2, 5, 1],
 [4, -3, 2, 6, -1, -3],
 [3, 1, 5, -8, 5, 1],
 [9, -2, -5, -2, 6, 1],
 [4, -1, 5, -7, -1, 7],
 [7, -6, 7, 1, -2, -1],
 [4, -2, -1, 8, -3, 1],
 [8, -7, 1, 5, -6, 3],
 [5, 4, -3, 2, -3, 1],
 [3, 5, -3, 2, -7, 2],
 [1, 2, -3, 4, 5, -6],
 [9, -8, 1, 2, 5, -1],
 [1, 5, 3, -7, -2, 8],
 [3, -1, 2, -4, 9, -6],
 [6, 2, -6, 5, -1, 2],
 [3, 6, -5, -4, 3, 6],
 [7, 1, -6, 2, -3, -1],
 [4, -1, -3, 7, -1, 2],
 [9, -7, 6, 1, -8, -1],
 [6, 1, -5, 2, -4, 7],
 [3, 1, -4, 9, -3, 1],
 [7, 1, -8, 5, 3, -7],
 [3, -1, 2, 5, -2, -7],
 [4, -2, -1, -1, 3, 1],
 [9, -4, 2, 2, -1, -7],
 [2, 7, -3, 1, -2, 3],
 [1, 7, 1, -6, 1, -4],
 [3, 1, -4, 1, 3, -1],
 [5, 4, -7, 5, 1, -8],
 [7, -1, 3, -1, -6, 1],
 [3, 1, -2, 6, -7, 2],
 [4, -2, -1, 7, -1, 2],
 [8, -5, 1, -4, 9, -5],
 [3, 1, -4, 6, -1, 2],
 [2, 6, 1, -4, 3, -6],
 [4, -1, 5, -3, -5, 9],
 [6, -5, 2, -3, 8, 1],
 [2, 5, -6, 7, 1, -4],
 [7, 2, -5, -2, -1, 5],
 [6, 2, -5, -1, 6, -7],
 [4, -1, -2, 8, -2, -5],
 [2, -1, 6, -7, 5, 4],
 [9, -2, -6, 2, 6, -2],
 [1, 5, 1, -2, -5, 6],
 [8, -7, 3, -4, 6, -1],
 [1, 8, -5, -3, 8, -9],
 [5, 4, -7, 2, -1, 5],
 [2, 7, -3, -5, 3, -4],
 [3, 5, -3, 1, -6, 1],
 [4, -1, 5, -1, -2, 3],
 [5, 2, -6, 3, -2, 7]]

MENTAL_LEVEL_SEQUENCES = {
    '5 yosh': MENTAL_5_6_YOSH_SEQUENCES,
    '6 yosh': MENTAL_5_6_YOSH_SEQUENCES,
    '8 yosh': MENTAL_8_9_YOSH_SEQUENCES,
    '9 yosh': MENTAL_8_9_YOSH_SEQUENCES,
}

MENTAL_ARITHMETIC_SEQUENCES = MENTAL_5_6_YOSH_SEQUENCES

def format_mental_number(value, is_first=False):
    return str(value)


def generate_mental_tasks(student, count=None):
    existing_tasks = list(MentalTask.objects.filter(student=student).order_by('task_order'))
    if existing_tasks:
        return existing_tasks

    created = []
    level_name = str(getattr(student.level, 'name', '') or '').strip()
    selected_sequences = MENTAL_LEVEL_SEQUENCES.get(level_name, MENTAL_ARITHMETIC_SEQUENCES)
    sequences = selected_sequences if count is None else selected_sequences[:count]

    for order, sequence in enumerate(sequences, start=1):
        flashes = [format_mental_number(value, index == 0) for index, value in enumerate(sequence)]
        correct_answer = sum(sequence)
        task = MentalTask.objects.create(
            student=student,
            task_order=order,
            flashes=flashes,
            expression=' '.join(flashes),
            correct_answer=correct_answer,
        )
        created.append(task)
    return created





from .serializers import (
    SubjectSerializer,
    LevelSerializer,
    StudentSerializer,
    QuestionAdminSerializer,
    QuestionForExamSerializer,
    ResultSerializer,
)


class IsAdminUser(permissions.BasePermission):
    def has_permission(self, request, view):
        return bool(request.user and request.user.is_authenticated and request.user.is_staff)


def get_user_profile(user):
    return getattr(user, 'admin_profile', None) if user and user.is_authenticated else None


def get_user_center(user):
    profile = get_user_profile(user)
    return getattr(profile, 'center', None) if profile else None


def is_main_admin(user):
    profile = get_user_profile(user)
    return bool(user and user.is_authenticated and user.is_staff and (user.is_superuser or not profile or not profile.center_id))


def can_manage_students(user):
    return bool(user and user.is_authenticated and user.is_staff)


def get_default_center():
    center, _ = Center.objects.get_or_create(name='LIDER.Uz Onlayn Olimpiada')
    return center


def get_center_by_name(name):
    center_name = str(name or '').strip()
    if not center_name:
        return get_default_center()
    center, _ = Center.objects.get_or_create(name=center_name)
    return center


class SubjectViewSet(viewsets.ModelViewSet):
    queryset = Subject.objects.all()
    serializer_class = SubjectSerializer
    permission_classes = [IsAdminUser]


class LevelViewSet(viewsets.ModelViewSet):
    queryset = Level.objects.select_related('subject').all()
    serializer_class = LevelSerializer
    permission_classes = [IsAdminUser]

    def get_queryset(self):
        qs = super().get_queryset()
        subject_id = self.request.query_params.get('subject')
        if subject_id:
            qs = qs.filter(subject_id=subject_id)
        return qs


class StudentViewSet(viewsets.ModelViewSet):
    queryset = Student.objects.select_related('subject', 'level', 'center').all()
    serializer_class = StudentSerializer
    permission_classes = [IsAdminUser]

    def get_queryset(self):
        qs = super().get_queryset()
        q = self.request.query_params.get('q')
        status_value = self.request.query_params.get('status')
        subject = self.request.query_params.get('subject')
        level = self.request.query_params.get('level')
        if q:
            qs = qs.filter(Q(first_name__icontains=q) | Q(last_name__icontains=q) | Q(code__icontains=q))
        if status_value:
            qs = qs.filter(status=status_value)
        if subject:
            qs = qs.filter(subject_id=subject)
        if level:
            qs = qs.filter(level_id=level)
        return qs.order_by('-created_at')

    def create(self, request, *args, **kwargs):
        if not can_manage_students(request.user):
            return Response({'detail': 'Sizga o‘quvchi yaratishga ruxsat yo‘q.'}, status=status.HTTP_403_FORBIDDEN)
        return super().create(request, *args, **kwargs)

    def perform_create(self, serializer):
        center = get_center_by_name(self.request.data.get('center_name'))
        serializer.save(center=center, branch='Boshqa')

    def update(self, request, *args, **kwargs):
        if not can_manage_students(request.user):
            return Response({'detail': 'Sizga o‘quvchini tahrirlashga ruxsat yo‘q.'}, status=status.HTTP_403_FORBIDDEN)
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        if not can_manage_students(request.user):
            return Response({'detail': 'Sizga o‘quvchini tahrirlashga ruxsat yo‘q.'}, status=status.HTTP_403_FORBIDDEN)
        return super().partial_update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        if not can_manage_students(request.user):
            return Response({'detail': 'Sizga o‘quvchini o‘chirishga ruxsat yo‘q.'}, status=status.HTTP_403_FORBIDDEN)
        return super().destroy(request, *args, **kwargs)

    @action(detail=False, methods=['post'], url_path='bulk-delete')
    def bulk_delete(self, request):
        if not can_manage_students(request.user):
            return Response({'detail': 'Sizga o‘quvchilarni o‘chirishga ruxsat yo‘q.'}, status=status.HTTP_403_FORBIDDEN)

        ids = request.data.get('ids') or []
        if not isinstance(ids, list):
            return Response({'detail': 'ids ro‘yxat ko‘rinishida yuborilishi kerak.'}, status=status.HTTP_400_BAD_REQUEST)

        cleaned_ids = []
        for value in ids:
            try:
                cleaned_ids.append(int(value))
            except (TypeError, ValueError):
                continue

        cleaned_ids = list(dict.fromkeys(cleaned_ids))
        if not cleaned_ids:
            return Response({'detail': 'O‘chirish uchun o‘quvchilar belgilanmagan.'}, status=status.HTTP_400_BAD_REQUEST)

        queryset = Student.objects.filter(id__in=cleaned_ids)
        deleted_count = queryset.count()
        queryset.delete()

        return Response({
            'deleted_count': deleted_count,
            'message': f'{deleted_count} ta o‘quvchi o‘chirildi.'
        })

    @action(detail=False, methods=['post'], url_path='import-excel')
    def import_excel(self, request):
        if not can_manage_students(request.user):
            return Response({'detail': 'Sizga Excel orqali o‘quvchi yaratishga ruxsat yo‘q.'}, status=status.HTTP_403_FORBIDDEN)
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'detail': 'Excel file yuborilmadi.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = load_workbook(file_obj)
            ws = wb.active
        except Exception:
            return Response({'detail': 'Excel faylni o‘qib bo‘lmadi.'}, status=status.HTTP_400_BAD_REQUEST)

        def normalize_header(value):
            text = str(value or '').strip().lower()
            replacements = {
                '‘': "'",
                '’': "'",
                '`': "'",
                'ʼ': "'",
                '№': 'no',
                '#': 'no',
            }
            for old, new_value in replacements.items():
                text = text.replace(old, new_value)
            text = ' '.join(text.split())
            return text

        headers = [normalize_header(cell.value) for cell in ws[1]]
        header_map = {name: idx for idx, name in enumerate(headers) if name}

        def cell_value(row, names):
            for name in names:
                idx = header_map.get(normalize_header(name))
                if idx is not None and idx < len(row):
                    value = row[idx]
                    return str(value).strip() if value is not None else ''
            return ''

        def split_full_name(full_name):
            parts = str(full_name or '').strip().split()
            if not parts:
                return '', ''
            if len(parts) == 1:
                return parts[0], ''
            return parts[0], ' '.join(parts[1:])

        def normalize_subject_name(value):
            raw = str(value or '').strip()
            lowered = raw.lower().replace('‘', "'").replace('’', "'")
            lowered = ' '.join(lowered.split())
            if lowered in ['ingliz', 'ingliz tili', 'english']:
                return 'Ingliz tili'
            if 'mental' in lowered:
                return 'Mental arifmetika'
            return raw

        def normalize_level_name(subject_name, value):
            raw = str(value or '').strip()
            lowered = raw.lower().replace('‘', "'").replace('’', "'")
            lowered = ' '.join(lowered.replace('_', '-').split())
            digits = ''.join(ch for ch in lowered if ch.isdigit())
            if subject_name == 'Ingliz tili':
                if lowered in ['qiyin', 'hard', 'difficult']:
                    return 'Qiyin'
                if digits in ['0', '1', '2', '3', '4', '5', '6', '7']:
                    return f'{digits}-sinf'
            if subject_name == 'Rus tili':
                if lowered in ['boxcha', 'bogcha', 'boqcha', 'maktabgacha']:
                    return 'boxcha'
                if digits in ['2', '3', '4', '5', '6', '7', '8', '9', '10']:
                    return f'{digits}-sinf'
            if subject_name == 'Matematika':
                if digits in ['0', '1', '2', '3', '4', '6', '8', '10', '11']:
                    return f'{digits}-sinf'
            if subject_name == 'Mental arifmetika':
                if '5' in digits or '5 yosh' in lowered:
                    return '5 yosh'
                if '6' in digits or '6 yosh' in lowered:
                    return '6 yosh'
                if '8' in digits or '8 yosh' in lowered:
                    return '8 yosh'
                if '9' in digits or '9 yosh' in lowered:
                    return '9 yosh'
            return raw

        created = []
        errors = []

        with transaction.atomic():
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                full_name = cell_value(row, [
                    'Ism familya',
                    'Ism familyasi',
                    'Ism Familyasi',
                    'F.I.Sh',
                    'FISH',
                    'FIO',
                    'Oquvchi',
                    "O'quvchi",
                    'O‘quvchi',
                ])
                first_name = cell_value(row, ['Ism', 'first_name', 'First name'])
                last_name = cell_value(row, ['Familya', 'last_name', 'Last name'])

                if full_name and not first_name and not last_name:
                    first_name, last_name = split_full_name(full_name)

                subject_name = normalize_subject_name(cell_value(row, ['Fan', 'Subject']))
                level_name = normalize_level_name(subject_name, cell_value(row, ['Sinfi', 'Sinf', 'Class', 'Grade', 'Daraja', 'Level']))
                branch_name = 'Boshqa'
                center_name = cell_value(row, [
                    'Oquv markaz nomi',
                    "O'quv markaz nomi",
                    'O‘quv markaz nomi',
                    'O‘quv markazi',
                    'Oquv markazi',
                    'Markaz nomi',
                    'Markaz',
                    'Center',
                    'Learning center',
                ])
                center = get_center_by_name(center_name)

                if not any([full_name, first_name, last_name, subject_name, level_name, center_name]):
                    continue

                if not all([first_name, subject_name, level_name]):
                    errors.append({
                        'row': row_num,
                        'error': "Majburiy ustunlar: Ism familya, Fan, Sinfi.",
                    })
                    continue

                try:
                    subject = Subject.objects.get(name__iexact=subject_name)
                except Subject.DoesNotExist:
                    errors.append({
                        'row': row_num,
                        'error': f'Fan topilmadi: {subject_name}. Avval shu fanni testlar bo‘limida yarating.',
                    })
                    continue

                try:
                    level = Level.objects.get(subject=subject, name__iexact=level_name)
                except Level.DoesNotExist:
                    errors.append({
                        'row': row_num,
                        'error': f'{subject.name} uchun sinf topilmadi: {level_name}.',
                    })
                    continue

                student = Student.objects.create(
                    first_name=first_name,
                    last_name=last_name,
                    subject=subject,
                    level=level,
                    center=center,
                    branch=branch_name,
                )
                created.append(student)

        return Response({
            'created_count': len(created),
            'errors': errors,
            'students': StudentSerializer(created, many=True).data,
        }, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        students = self.get_queryset().select_related('subject', 'level', 'center', 'result')

        wb = Workbook()
        ws = wb.active
        ws.title = 'Barcha oquvchilar'

        headers = ['№', 'Ism familya', "O'quv markaz nomi", 'Fan', 'Sinfi', 'Code', 'Status', 'Nechta to‘g‘ri', 'Jami savollar', 'Foiz']
        ws.append(headers)

        header_fill = PatternFill('solid', fgColor='1F4E79')
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        status_labels = {
            Student.Status.NOT_STARTED: 'Ishlamagan',
            Student.Status.IN_PROGRESS: 'Ishlayapti',
            Student.Status.COMPLETED: 'Ishlab bo‘ldi',
        }

        for idx, student in enumerate(students, start=1):
            try:
                result = student.result
            except ObjectDoesNotExist:
                result = None

            correct_count = result.correct_count if result else ''
            total_questions = result.total_questions if result else ''
            percent = f'{result.percent:.1f}%' if result else ''

            ws.append([
                idx,
                student.full_name,
                student.center.name if student.center else '',
                student.subject.name,
                student.level.name,
                student.code,
                status_labels.get(student.status, student.status),
                correct_count,
                total_questions,
                percent,
            ])

        ws.freeze_panes = 'A2'
        for column_cells in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max_length + 4, 42)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="barcha_oquvchilar.xlsx"'
        return response

class QuestionViewSet(viewsets.ModelViewSet):
    queryset = Question.objects.select_related('subject', 'level').all()
    serializer_class = QuestionAdminSerializer
    permission_classes = [IsAdminUser]

    def get_queryset(self):
        qs = super().get_queryset()
        subject = self.request.query_params.get('subject')
        level = self.request.query_params.get('level')
        if subject:
            qs = qs.filter(subject_id=subject)
        if level:
            qs = qs.filter(level_id=level)
        return qs.order_by('subject__name', 'level__name', 'id')

    def create(self, request, *args, **kwargs):
        if not is_main_admin(request.user):
            return Response({'detail': 'Sizga test qo‘shishga ruxsat yo‘q.'}, status=status.HTTP_403_FORBIDDEN)
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        if not is_main_admin(request.user):
            return Response({'detail': 'Sizga testni tahrirlashga ruxsat yo‘q.'}, status=status.HTTP_403_FORBIDDEN)
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        if not is_main_admin(request.user):
            return Response({'detail': 'Sizga testni tahrirlashga ruxsat yo‘q.'}, status=status.HTTP_403_FORBIDDEN)
        return super().partial_update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        if not is_main_admin(request.user):
            return Response({'detail': 'Sizga testni o‘chirishga ruxsat yo‘q.'}, status=status.HTTP_403_FORBIDDEN)
        return super().destroy(request, *args, **kwargs)


class ResultViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Result.objects.select_related('student', 'student__subject', 'student__level', 'student__center').prefetch_related('answers', 'mental_answers').all()
    serializer_class = ResultSerializer
    permission_classes = [IsAdminUser]

    def get_queryset(self):
        qs = super().get_queryset()
        center = self.request.query_params.get('center')
        user_center = get_user_center(self.request.user)

        if not is_main_admin(self.request.user) and user_center:
            qs = qs.filter(student__center=user_center)
        elif center:
            qs = qs.filter(student__center_id=center)

        return qs.order_by('-created_at')

    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Natijalar'

        headers = [
            '№', 'Ism Familya', 'Fan', 'Sinfi', 'Status code',
            'Nechta to‘g‘ri', 'Jami savollar', 'Foiz', 'Boshlangan vaqti',
            'Tugatgan vaqti', 'Sarflagan vaqt'
        ]
        ws.append(headers)

        header_fill = PatternFill('solid', fgColor='1F4E79')
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        for idx, result in enumerate(self.get_queryset(), start=1):
            spent = str(timedelta(seconds=result.spent_seconds))
            ws.append([
                idx,
                result.student.full_name,
                result.student.subject.name,
                result.student.level.name,
                result.student.code,
                result.correct_count,
                result.total_questions,
                f'{result.percent:.1f}%',
                timezone.localtime(result.started_at).strftime('%Y-%m-%d %H:%M:%S') if result.started_at else '',
                timezone.localtime(result.finished_at).strftime('%Y-%m-%d %H:%M:%S') if result.finished_at else '',
                spent,
            ])

        for column_cells in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max_length + 3, 40)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="olimpiada_natijalari.xlsx"'
        return response


    @action(detail=False, methods=['get'], url_path='mental-answers')
    def mental_answers(self, request):
        qs = self.get_queryset().filter(mental_answers__isnull=False).distinct().prefetch_related('mental_answers')
        serializer = self.get_serializer(qs, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'], url_path='mental-answers-export')
    def mental_answers_export(self, request):
        qs = self.get_queryset().filter(mental_answers__isnull=False).distinct().prefetch_related('mental_answers')

        wb = Workbook()
        ws = wb.active
        ws.title = 'Mental javoblari'

        headers = [
            '№', 'Ism Familya', 'Fan', 'Sinfi', 'Status code',
            'Nechta to‘g‘ri', 'Jami savollar', 'Foiz', 'Sarflagan vaqt', 'Misol №', 'Misol', "O'quvchi javobi",
            "To'g'ri javob", 'Holat'
        ]
        ws.append(headers)

        header_fill = PatternFill('solid', fgColor='1F4E79')
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        row_index = 1
        for result in qs:
            tasks = list(result.mental_answers.all().order_by('task_order'))
            if not tasks:
                continue
            spent = str(timedelta(seconds=result.spent_seconds))
            for task in tasks:
                ws.append([
                    row_index,
                    result.student.full_name,
                    result.student.subject.name,
                    result.student.level.name,
                    result.student.code,
                    result.correct_count,
                    result.total_questions,
                    f'{result.percent:.1f}%',
                    spent,
                    task.task_order,
                    task.expression,
                    task.student_answer if task.student_answer is not None else '',
                    task.correct_answer,
                    "To'g'ri" if task.is_correct else "Noto'g'ri",
                ])
            row_index += 1

        for column_cells in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max_length + 3, 45)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="mental_javoblari.xlsx"'
        return response


class PublicResultLookupAPIView(APIView):
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        code = str(request.data.get('code', '')).strip()
        if not code:
            return Response({'detail': 'Status code kiriting.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            student = Student.objects.select_related('subject', 'level', 'center').get(code=code)
        except Student.DoesNotExist:
            return Response({'detail': 'Bunday code topilmadi.'}, status=status.HTTP_404_NOT_FOUND)

        if student.status != Student.Status.COMPLETED:
            return Response({'detail': 'Bu o‘quvchi hali testni yakunlamagan.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            result = student.result
        except ObjectDoesNotExist:
            return Response({'detail': 'Bu code uchun natija topilmadi.'}, status=status.HTTP_404_NOT_FOUND)

        is_mental = is_mental_subject(student.subject.name)
        mental_answers = []
        if is_mental:
            mental_answers = [
                {
                    'id': task.id,
                    'task_order': task.task_order,
                    'expression': task.expression,
                    'correct_answer': task.correct_answer,
                    'student_answer': task.student_answer,
                    'is_correct': task.is_correct,
                }
                for task in result.mental_answers.all().order_by('task_order')
            ]

        return Response({
            'student_full_name': student.full_name,
            'student_code': student.code,
            'subject_name': student.subject.name,
            'level_name': student.level.name,
            'center_name': student.center.name,
            'total_questions': result.total_questions,
            'correct_count': result.correct_count,
            'percent': result.percent,
            'spent_seconds': result.spent_seconds,
            'finished_at': result.finished_at,
            'is_mental': is_mental,
            'mental_answers': mental_answers,
        })


class PublicResultsListAPIView(APIView):
    permission_classes = [permissions.AllowAny]

    def get(self, request):
        center_id = str(request.query_params.get('center', '') or '').strip()
        subject_id = str(request.query_params.get('subject', '') or '').strip()

        base_qs = Result.objects.select_related(
            'student', 'student__subject', 'student__level', 'student__center'
        ).filter(student__status=Student.Status.COMPLETED)

        qs = base_qs
        if center_id:
            qs = qs.filter(student__center_id=center_id)
        if subject_id:
            qs = qs.filter(student__subject_id=subject_id)

        results = []
        for result in qs.order_by('-finished_at', '-created_at'):
            student = result.student
            results.append({
                'id': result.id,
                'student_full_name': student.full_name,
                'center_id': student.center_id,
                'center_name': student.center.name if student.center else '',
                'subject_id': student.subject_id,
                'subject_name': student.subject.name,
                'level_name': student.level.name,
                'correct_count': result.correct_count,
                'total_questions': result.total_questions,
                'percent': result.percent,
                'spent_seconds': result.spent_seconds,
                'finished_at': result.finished_at,
            })

        centers = [
            {'id': row['student__center_id'], 'name': row['student__center__name']}
            for row in base_qs.exclude(student__center__isnull=True)
            .values('student__center_id', 'student__center__name')
            .distinct()
            .order_by('student__center__name')
        ]
        subjects = [
            {'id': row['student__subject_id'], 'name': row['student__subject__name']}
            for row in base_qs.values('student__subject_id', 'student__subject__name')
            .distinct()
            .order_by('student__subject__name')
        ]

        return Response({
            'results': results,
            'centers': centers,
            'subjects': subjects,
        })

class ExamStartAPIView(APIView):
    permission_classes = [permissions.AllowAny]

    @transaction.atomic
    def post(self, request):
        code = str(request.data.get('code', '')).strip()
        if not code:
            return Response({'detail': 'Status code kiriting.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            student = Student.objects.select_related('subject', 'level', 'center').select_for_update().get(code=code)
        except Student.DoesNotExist:
            return Response({'detail': 'Bunday code topilmadi.'}, status=status.HTTP_404_NOT_FOUND)

        if student.status == Student.Status.COMPLETED or student.is_used:
            return Response({'detail': 'Bu code oldin ishlatilgan.'}, status=status.HTTP_400_BAD_REQUEST)

        is_mental_exam = is_mental_subject(student.subject.name)

        # Test birinchi marta boshlanganda vaqt shu paytdan hisoblanadi.
        # Agar o'quvchi bilmastan chiqib ketib qayta kirsa, started_at o'zgarmaydi:
        # mental arifmetika o'sha qolgan vaqt va keyingi javob berilmagan savoldan davom etadi.
        now = timezone.now()
        if not student.started_at:
            student.started_at = now
        student.status = Student.Status.IN_PROGRESS
        student.save(update_fields=['status', 'started_at'])

        remaining_seconds = get_remaining_seconds(student, now)

        if is_mental_exam:
            mental_tasks = generate_mental_tasks(student)
            return Response({
                'mode': 'mental',
                'student': StudentSerializer(student).data,
                'duration_minutes': get_exam_duration_minutes(student),
                'started_at': student.started_at,
                'server_now': now,
                'remaining_seconds': remaining_seconds,
                'mental_tasks': [
                    {
                        'id': task.id,
                        'task_order': task.task_order,
                        'flashes': task.flashes,
                        'task_display_ms': 3000,
                        'student_answer': task.student_answer,
                    }
                    for task in mental_tasks
                ],
            })

        questions = Question.objects.filter(subject=student.subject, level=student.level).order_by('id')
        if not questions.exists():
            student.status = Student.Status.NOT_STARTED
            student.started_at = None
            student.save(update_fields=['status', 'started_at'])
            return Response({'detail': 'Bu fan va sinf uchun testlar hali qo‘shilmagan.'}, status=status.HTTP_400_BAD_REQUEST)

        return Response({
            'mode': 'test',
            'student': StudentSerializer(student).data,
            'duration_minutes': get_exam_duration_minutes(student),
            'started_at': student.started_at,
            'server_now': now,
            'remaining_seconds': remaining_seconds,
            'questions': QuestionForExamSerializer(questions, many=True).data,
        })



class MentalProgressAPIView(APIView):
    permission_classes = [permissions.AllowAny]

    @transaction.atomic
    def post(self, request):
        code = str(request.data.get('code', '')).strip()
        task_id = request.data.get('task_id')
        raw_answer = request.data.get('answer', '')

        if not code:
            return Response({'detail': 'Status code yuborilmadi.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            student = Student.objects.select_related('subject', 'level', 'center').select_for_update().get(code=code)
        except Student.DoesNotExist:
            return Response({'detail': 'Bunday code topilmadi.'}, status=status.HTTP_404_NOT_FOUND)

        if student.status == Student.Status.COMPLETED or student.is_used:
            return Response({'detail': 'Bu code oldin ishlatilgan.'}, status=status.HTTP_400_BAD_REQUEST)

        if not is_mental_subject(student.subject.name):
            return Response({'detail': 'Bu endpoint faqat mental arifmetika uchun.'}, status=status.HTTP_400_BAD_REQUEST)

        if not student.started_at:
            student.started_at = timezone.now()
            student.status = Student.Status.IN_PROGRESS
            student.save(update_fields=['started_at', 'status'])

        try:
            task = MentalTask.objects.select_for_update().get(student=student, id=int(task_id))
        except (MentalTask.DoesNotExist, TypeError, ValueError):
            return Response({'detail': 'Mental savol topilmadi.'}, status=status.HTTP_404_NOT_FOUND)

        try:
            answer = int(str(raw_answer).strip())
        except (TypeError, ValueError):
            return Response({'detail': 'Javob son ko‘rinishida bo‘lishi kerak.'}, status=status.HTTP_400_BAD_REQUEST)

        task.student_answer = answer
        task.is_correct = answer == task.correct_answer
        task.save(update_fields=['student_answer', 'is_correct'])

        return Response({
            'id': task.id,
            'task_order': task.task_order,
            'student_answer': task.student_answer,
            'is_correct': task.is_correct,
        })


class ExamSubmitAPIView(APIView):
    permission_classes = [permissions.AllowAny]

    def submit_mental(self, student, answers):
        task_ids = [item.get('task_id') for item in answers if isinstance(item, dict)]
        tasks = list(MentalTask.objects.select_for_update().filter(student=student, id__in=task_ids).order_by('task_order'))

        # Frontend yuborgan barcha ko‘rsatilgan misollar tekshiriladi.
        # Javob kiritilmagan misol noto‘g‘ri deb saqlanadi.
        answer_map = {}
        for item in answers:
            if not isinstance(item, dict):
                continue
            try:
                task_id = int(item.get('task_id'))
                answer = int(str(item.get('answer', '')).strip())
            except Exception:
                continue
            answer_map[task_id] = answer

        finished_at = timezone.now()
        started_at = student.started_at or finished_at
        spent_seconds = calculate_spent_seconds_from_request(self.request, student, finished_at)

        result = Result.objects.create(
            student=student,
            total_questions=len(tasks),
            correct_count=0,
            percent=0,
            started_at=started_at,
            finished_at=finished_at,
            spent_seconds=spent_seconds,
        )

        correct_count = 0
        for task in tasks:
            student_answer = answer_map.get(task.id)
            is_correct = student_answer == task.correct_answer
            if is_correct:
                correct_count += 1
            task.result = result
            task.student_answer = student_answer
            task.is_correct = is_correct
            task.save(update_fields=['result', 'student_answer', 'is_correct'])

        percent = (correct_count / len(tasks) * 100) if tasks else 0
        result.correct_count = correct_count
        result.percent = round(percent, 2)
        result.save(update_fields=['correct_count', 'percent'])

        student.status = Student.Status.COMPLETED
        student.finished_at = finished_at
        student.is_used = True
        student.save(update_fields=['status', 'finished_at', 'is_used'])

        return Response(ResultSerializer(result).data, status=status.HTTP_201_CREATED)

    @transaction.atomic
    def post(self, request):
        code = str(request.data.get('code', '')).strip()
        answers = request.data.get('answers', [])

        if not code:
            return Response({'detail': 'Status code yuborilmadi.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            student = Student.objects.select_related('subject', 'level', 'center').select_for_update().get(code=code)
        except Student.DoesNotExist:
            return Response({'detail': 'Bunday code topilmadi.'}, status=status.HTTP_404_NOT_FOUND)

        if student.status == Student.Status.COMPLETED or student.is_used:
            return Response({'detail': 'Bu code oldin ishlatilgan.'}, status=status.HTTP_400_BAD_REQUEST)

        if student.status != Student.Status.IN_PROGRESS:
            return Response({'detail': 'Avval testni boshlash kerak.'}, status=status.HTTP_400_BAD_REQUEST)

        if is_mental_subject(student.subject.name):
            return self.submit_mental(student, answers)

        answer_map = {}
        for item in answers:
            try:
                qid = int(item.get('question_id'))
            except Exception:
                continue
            selected = str(item.get('answer', '')).strip().upper()
            if selected in ['A', 'B', 'C', 'D']:
                answer_map[qid] = selected

        questions = list(Question.objects.filter(subject=student.subject, level=student.level).order_by('id'))
        correct_count = 0
        finished_at = timezone.now()
        started_at = student.started_at or finished_at
        spent_seconds = calculate_spent_seconds_from_request(self.request, student, finished_at)

        result = Result.objects.create(
            student=student,
            total_questions=len(questions),
            correct_count=0,
            percent=0,
            started_at=started_at,
            finished_at=finished_at,
            spent_seconds=spent_seconds,
        )

        answer_objects = []
        for question in questions:
            selected = answer_map.get(question.id, '')
            is_correct = selected == question.correct_answer
            if is_correct:
                correct_count += 1
            answer_objects.append(StudentAnswer(
                result=result,
                question=question,
                selected_answer=selected,
                is_correct=is_correct,
            ))
        StudentAnswer.objects.bulk_create(answer_objects)

        percent = (correct_count / len(questions) * 100) if questions else 0
        result.correct_count = correct_count
        result.percent = round(percent, 2)
        result.save(update_fields=['correct_count', 'percent'])

        student.status = Student.Status.COMPLETED
        student.finished_at = finished_at
        student.is_used = True
        student.save(update_fields=['status', 'finished_at', 'is_used'])

        return Response(ResultSerializer(result).data, status=status.HTTP_201_CREATED)
